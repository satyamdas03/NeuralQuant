"""QuantFactor Engine — Excel-based data ingestor.

Reads the QuantFactor Excel workbook (US_Stock_Analysis_Coloured.xlsx) which contains
pre-computed quintile scores and raw fundamental data for SP500, SmallMidCap,
and NSE 100 universes. This replaces the slow yfinance-based collector for
nightly updates.

The Excel is the source of truth — it's updated by the QuantFactor pipeline with fresh
data, and we read it directly into Supabase.

Usage:
    df = read_anjali_excel("path/to/US_Stock_Analysis_Coloured.xlsx")
    # Or per-sheet:
    df_sp500 = read_anjali_sheet(path, sheet="S&P 500 Analysis", market="US")
    df_smc = read_anjali_sheet(path, sheet="SmallMidCap Analysis", market="US")
    df_nse = read_anjali_sheet(path, sheet="NSE 100 Analysis", market="IN")

For Supabase direct ingestion (CI workflow):
    ingestor = ExcelIngestor(supabase_client)
    stats = ingestor.ingest("path/to/US_Stock_Analysis_Coloured.xlsx")
"""
from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Literal

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from supabase import Client

from .mappings import SHEET_META, INDEX_GROUP_MAP, _SP500_COL_MAP, _NSE_COL_MAP
from .scorer import hex_to_quintile

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column mapping: Excel column → our DB column (pandas-based path)
# ---------------------------------------------------------------------------

# SP500 / SmallMidCap sheet columns (raw values only — quintiles come from cell colors)
_PANDAS_SP500_COL_MAP = {
    "Ticker": "ticker",
    "Sector": "sector",
    "Sub Sector": "sub_sector",
    "Sales YoY Growth": "sales_yoy_growth",
    "NetProfit YoY Growth": "net_profit_yoy_growth",
    "Sales TTM 1Yr Growth": "sales_ttm_growth",
    "NetProfit TTM 1Yr Growth": "net_profit_ttm_growth",
    "QoQ Sales Growth": "qoq_sales_growth",
    "QoQ Profit Growth": "qoq_profit_growth",
    "3M Return": "return_3m",
    "6M Return": "return_6m",
    "1Yr Return": "return_1yr",
    "2Yr Return": "return_2yr",
    "PE Ratio": "pe_ratio",
    "Future PE": "future_pe",
    "TTM PEG": "ttm_peg",
    "Future PEG": "future_peg",
    "PB Ratio": "pb_ratio",
    "EV/Sales": "ev_sales",
    "EV/EBITDA": "ev_ebitda",
    "Market Cap (B)": "market_cap_bn",
    "Revenue (B)": "revenue_bn",
    "TTM Revenue (B)": "ttm_revenue_bn",
    "QtrStd": "qtr_std",
    "YrStd": "yr_std",
    "Qtr Beta": "qtr_beta",
    "Yr Beta": "yr_beta",
    "DII Quarter": "dii_quarter",
    "DII 1Yr": "dii_1yr",
    "FII Quarter": "fii_quarter",
    "FII 1Yr": "fii_1yr",
    "RETURN SCORE": "return_score",
    "GROWTH SCORE": "growth_score",
    "VALUATION SCORE": "valuation_score",
    "RISK SCORE": "risk_score",
}

# NSE sheet has different column order + extra columns
_PANDAS_NSE_COL_MAP = {
    "NseCode": "ticker",
    "Sector": "sector",
    "Sub Sector": "sub_sector",
    "Sales YoY Growth": "sales_yoy_growth",
    "NetProfit YoY Growth": "net_profit_yoy_growth",
    "Sales TTM 1Yr Growth": "sales_ttm_growth",
    "NetProfit TTM 1Yr Growth": "net_profit_ttm_growth",
    # NSE doesn't have QoQ columns
    "3M Return": "return_3m",
    "6M Return": "return_6m",
    "1Yr Return": "return_1yr",
    "2Yr Return": "return_2yr",
    "PE Ratio": "pe_ratio",
    "Future PE": "future_pe",
    "TTM PEG": "ttm_peg",
    "Future PEG": "future_peg",
    # NSE extra columns
    "Alpha": "alpha",
    "Risk": "risk_score_nse",
    "Final Score": "final_score",
    "DII Quarter": "dii_quarter",
    "DII 1Yr": "dii_1yr",
    "FII Quarter": "fii_quarter",
    "FII 1Yr": "fii_1yr",
    "QtrStd": "qtr_std",
    "YrStd": "yr_std",
    "Qtr Beta": "qtr_beta",
    "Yr Beta": "yr_beta",
    "RETURN SCORE": "return_score",
    "GROWTH SCORE": "growth_score",
    "VALUATION SCORE": "valuation_score",
    "RISK SCORE": "risk_score",
    # Strategy columns (mostly None)
    "Rebalance Date": "rebalance_date",
    "Future Return": "future_return",
    "Strategy Stocks": "strategy_stocks",
    "Stocks List": "stocks_list",
}


def _sanitize_df(df: pd.DataFrame) -> pd.DataFrame:
    """Clean DataFrame: replace inf with None, convert numpy types."""
    df = df.copy()
    for col in df.columns:
        df[col] = df[col].replace([np.inf, -np.inf], None)
        if df[col].dtype in [np.float64, np.float32, np.int64, np.int32]:
            df[col] = df[col].where(pd.notna(df[col]), None)
    return df


def read_anjali_sheet(
    path: str | Path,
    sheet: str = "S&P 500 Analysis",
    market: Literal["US", "IN"] = "US",
) -> pd.DataFrame:
    """Read a single sheet from the QuantFactor Excel workbook.

    Args:
        path: Path to US_Stock_Analysis_Coloured.xlsx
        sheet: Sheet name (default "S&P 500 Analysis")
        market: "US" or "IN"

    Returns:
        DataFrame with our DB column names, plus market and computed fields.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"QuantFactor Excel not found: {path}")

    raw = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
    logger.info(f"Read {len(raw)} rows from '{sheet}' sheet")

    col_map = _PANDAS_NSE_COL_MAP if market == "IN" else _PANDAS_SP500_COL_MAP
    df = raw.rename(columns=col_map)

    # Drop columns not in our map (like Index Name for NSE)
    keep_cols = [c for c in df.columns if c in col_map.values()]
    df = df[keep_cols].copy()

    # Add market
    df["market"] = market

    # Compute loss flags from net_profit columns
    df["loss_profit_yoy"] = df.get("net_profit_yoy_growth").apply(
        lambda x: True if pd.notna(x) and x < 0 else False
    )
    df["loss_profit_ttm"] = df.get("net_profit_ttm_growth").apply(
        lambda x: True if pd.notna(x) and x < 0 else False
    )
    if "qoq_profit_growth" in df.columns:
        df["loss_profit_qoq"] = df["qoq_profit_growth"].apply(
            lambda x: True if pd.notna(x) and x < 0 else False
        )
    else:
        df["loss_profit_qoq"] = False

    # Compute composite QuantFactor score (sum of 4 quintile scores, -16 to +16)
    score_cols = ["return_score", "growth_score", "valuation_score", "risk_score"]
    if all(c in df.columns for c in score_cols):
        df["composite_anjali_score"] = df[score_cols].sum(axis=1)
    else:
        df["composite_anjali_score"] = None

    # Add index_group (universe name)
    if sheet == "S&P 500 Analysis":
        df["index_group"] = "SP500"
    elif sheet == "SmallMidCap Analysis":
        # Split by Index column if present
        if "Index" in raw.columns:
            idx_map = {"MidCap 400": "SP400", "SmallCap 600": "SP600"}
            df["index_group"] = raw["Index"].map(idx_map).fillna("SP400+SP600")
        else:
            df["index_group"] = "SP400+SP600"
    elif sheet == "NSE 100 Analysis":
        df["index_group"] = "NIFTY100"
    else:
        df["index_group"] = None

    # Add data_collected_at timestamp
    df["data_collected_at"] = datetime.now(timezone.utc).isoformat()

    # For NSE: add .NS suffix to tickers if not present
    if market == "IN":
        df["ticker"] = df["ticker"].apply(
            lambda t: t + ".NS" if pd.notna(t) and "." not in str(t) else t
        )

    # PB Ratio: exclude negative values
    if "pb_ratio" in df.columns:
        df.loc[df["pb_ratio"] < 0, "pb_ratio"] = None

    # Sanitize: replace inf, convert numpy types
    df = _sanitize_df(df)

    logger.info(f"Processed {len(df)} rows for {market} market, composite scores: {df['composite_anjali_score'].notna().sum()}")
    return df


def read_anjali_excel(path: str | Path) -> dict[str, pd.DataFrame]:
    """Read all 3 sheets from the QuantFactor Excel workbook.

    Returns:
        Dict with keys 'sp500', 'smallmidcap', 'nse100' mapping to DataFrames.
    """
    path = Path(path)
    sheets = {
        "sp500": ("S&P 500 Analysis", "US"),
        "smallmidcap": ("SmallMidCap Analysis", "US"),
        "nse100": ("NSE 100 Analysis", "IN"),
    }
    result = {}
    for key, (sheet_name, market) in sheets.items():
        try:
            result[key] = read_anjali_sheet(path, sheet=sheet_name, market=market)
            logger.info(f"Loaded {key}: {len(result[key])} rows")
        except Exception as e:
            logger.warning(f"Failed to load sheet '{sheet_name}': {e}")
    return result


def ingest_excel_to_supabase(path: str | Path) -> dict[str, int]:
    """Read QuantFactor Excel and upsert all data to Supabase.

    Returns:
        Dict with counts per sheet.
    """
    from .ingestor import ingest_to_supabase

    sheet_markets = {"sp500": "US", "smallmidcap": "US", "nse100": "IN"}
    sheets = read_anjali_excel(path)
    counts = {}
    for key, df in sheets.items():
        if df.empty:
            continue
        count = ingest_to_supabase(df, market=sheet_markets[key])
        counts[key] = count
        logger.info(f"Ingested {key}: {count} rows")
    return counts


# ---------------------------------------------------------------------------
# ExcelIngestor — openpyxl-based Supabase ingestion (for CI workflow)
# Reads cell fill colors to extract quintile labels alongside raw values.
# ---------------------------------------------------------------------------

_BATCH_SIZE = 500


class ExcelIngestor:
    """Reads the QuantFactor Excel and upserts to Supabase anjali_enrichment table.

    Uses openpyxl to read both cell values AND fill colors (quintile labels),
    then performs per-market DELETE+INSERT refresh.
    """

    def __init__(self, supabase: Client):
        self.supabase = supabase

    def ingest(self, source: Path | str) -> dict[str, int]:
        """Parse workbook, build records, and refresh Supabase per market.

        Returns dict of sheet_name → row count ingested.
        """
        source = Path(source)
        logger.info("Loading workbook: %s", source)
        wb = load_workbook(str(source), data_only=True)

        us_records: list[dict[str, Any]] = []
        in_records: list[dict[str, Any]] = []
        stats: dict[str, int] = {}

        for sheet_name, meta in SHEET_META.items():
            if sheet_name not in wb.sheetnames:
                logger.warning("Sheet %s not found in workbook, skipping", sheet_name)
                continue
            ws = wb[sheet_name]
            records = self._parse_sheet(ws, meta)
            if meta["market"] == "US":
                us_records.extend(records)
            else:
                in_records.extend(records)
            stats[sheet_name] = len(records)
            logger.info("Sheet %s: %d records", sheet_name, len(records))

        if us_records:
            self._refresh_market("US", us_records)
            logger.info("Refreshed US market: %d rows", len(us_records))
        if in_records:
            self._refresh_market("IN", in_records)
            logger.info("Refreshed IN market: %d rows", len(in_records))

        wb.close()
        return stats

    def _parse_sheet(
        self, ws, meta: dict[str, Any]
    ) -> list[dict[str, Any]]:
        """Parse a single worksheet into a list of record dicts."""
        header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col_idx = {h: i for i, h in enumerate(header_row) if h is not None}

        col_map = meta["col_map"]
        derive_index = meta.get("index_group_derive")
        fixed_index_group = meta.get("index_group")
        market = meta["market"]
        fetched_at = datetime.now(timezone.utc).isoformat()

        records: list[dict[str, Any]] = []

        for row in ws.iter_rows(min_row=2):
            if row[0].value is None:
                continue

            record: dict[str, Any] = {
                "market": market,
                "fetched_at": fetched_at,
            }

            # Determine index_group
            if derive_index:
                idx_col = col_idx.get(derive_index)
                if idx_col is not None:
                    idx_val = row[idx_col].value
                    record["index_group"] = INDEX_GROUP_MAP.get(str(idx_val).strip())
                if not record.get("index_group"):
                    continue
            else:
                record["index_group"] = fixed_index_group

            # Map each column
            for excel_header, (db_col, quintile_col, dtype) in col_map.items():
                if excel_header not in col_idx:
                    continue
                cell = row[col_idx[excel_header]]
                raw = self._coerce(cell.value, dtype)
                record[db_col] = raw

                # Extract quintile from cell fill color
                if quintile_col:
                    color_hex = cell.fill.start_color.rgb
                    q = hex_to_quintile(color_hex)
                    record[quintile_col] = q

            # Skip rows without a ticker
            if not record.get("ticker"):
                continue

            # Remove internal _index_derive field if present
            record.pop("_index_derive", None)

            records.append(record)

        return records

    @staticmethod
    def _coerce(value: Any, dtype: str) -> Any:
        """Coerce an Excel cell value to the expected Python type."""
        if value is None or value == "":
            return None
        if dtype == "float":
            try:
                return float(value)
            except (ValueError, TypeError):
                return None
        if dtype == "text":
            return str(value).strip()
        return value

    def _refresh_market(self, market: str, records: list[dict[str, Any]]) -> None:
        """DELETE all rows for a market, then INSERT in batches."""
        logger.info("Deleting existing %s rows from anjali_enrichment", market)
        self.supabase.table("anjali_enrichment").delete().eq("market", market).execute()

        total = len(records)
        for i in range(0, total, _BATCH_SIZE):
            batch = records[i : i + _BATCH_SIZE]
            self.supabase.table("anjali_enrichment").insert(batch).execute()
            logger.info(
                "Inserted batch %d-%d of %d for market=%s",
                i + 1,
                min(i + _BATCH_SIZE, total),
                total,
                market,
            )